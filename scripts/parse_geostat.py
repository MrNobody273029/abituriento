"""
Parses GeoStat higher education Excel files and outputs structured JSON.
Source: https://www.geostat.ge/ka/modules/categories/61/umaghlesi-ganatleba
"""
import openpyxl, json, sys, os
sys.stdout.reconfigure(encoding="utf-8")

BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
RAW = os.path.join(BASE, "data_raw")
OUT = os.path.join(BASE, "data_structured")
os.makedirs(OUT, exist_ok=True)

FIELD_MAP = {
    "მიღებულთა რაოდენობა, სულ": "total",
    "განათლება": "education",
    "ჰუმანიტარული მეცნიერებები და ხელოვნება": "humanities",
    "სოციალური მეცნიერებები, ბიზნესი და სამართალი": "social_business_law",
    "მეცნიერება": "science",
    "საინჟინრო, დამამუშავებელი და სამშენებლო დარგები": "engineering",
    "სოფლის მეურნეობა": "agriculture",
    "ჯანდაცვა და სოციალური უზრუნველყოფა": "health",
    "მომსახურება": "services",
}

TARGET_YEARS = ["2022-2023", "2023-2024", "2024-2025", "2025-2026"]


def find_year_cols(ws):
    cols = {}
    for i, cell in enumerate(ws[2]):
        if cell.value:
            for y in TARGET_YEARS:
                if y in str(cell.value):
                    cols[i] = y
    return cols


def parse_sheet(ws):
    col_years = find_year_cols(ws)
    result = {y: {"total": {}, "state": {}, "private": {}, "female": {}, "male": {}} for y in TARGET_YEARS}
    section = "total"

    for row in ws.iter_rows(values_only=True):
        if not row[0]:
            continue
        label = str(row[0]).strip()
        if label in ["მათ შორის:", "მათ შორის"]:
            continue
        if "კაცი" in label and "მიღება" in label:
            continue

        if "მათ შორის ქალი" in label:
            section = "female"
            continue
        if "მათ შორის კაცი" in label:
            section = "male"
            continue
        if "სახელმწიფო" in label and "უმაღლეს" in label:
            section = "state"
            continue
        if "კერძო" in label and "უმაღლეს" in label:
            section = "private"
            continue
        if label.startswith("მიღებულთა რაოდენობა") and section not in ("state", "private", "female", "male"):
            section = "total"

        field_key = None
        for ka, en in FIELD_MAP.items():
            if label.startswith(ka.strip()):
                field_key = en
                break
        if not field_key:
            continue

        for col_i, year in col_years.items():
            val = row[col_i]
            if val is None:
                continue
            try:
                val = int(str(val).replace(" ", "").replace("\xa0", ""))
            except:
                continue
            result[year][section][field_key] = val

    return result


wb1 = openpyxl.load_workbook(os.path.join(RAW, "01_admissions_by_program.xlsx"))
admissions = parse_sheet(wb1["მიღება"])
enrollment = parse_sheet(wb1["რიცხოვნობა"])
graduates = parse_sheet(wb1["კურსდამთავრებულები"])

output = {
    "source": "GeoStat — Statistics Georgia",
    "url": "https://www.geostat.ge/ka/modules/categories/61/umaghlesi-ganatleba",
    "years": TARGET_YEARS,
    "fields": list(FIELD_MAP.values()),
    "admissions": admissions,
    "enrollment": enrollment,
    "graduates": graduates,
}

out_path = os.path.join(OUT, "geostat_higher_education.json")
with open(out_path, "w", encoding="utf-8") as f:
    json.dump(output, f, ensure_ascii=False, indent=2)

print(f"Saved: {out_path}")
print("\nSample — 2024-2025 total admissions:")
for field, val in admissions["2024-2025"]["total"].items():
    print(f"  {field}: {val}")
