"""
Parse GeoStat higher education data - admissions + graduates
Outputs clean JSON with only verified real data.
"""
import openpyxl, json, sys, os
sys.stdout.reconfigure(encoding="utf-8")

BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
RAW = os.path.join(BASE, "data_raw")
OUT = os.path.join(BASE, "data_structured")
os.makedirs(OUT, exist_ok=True)

TARGET_YEARS = ["2022-2023", "2023-2024", "2024-2025", "2025-2026"]

FIELD_MAP = {
    "განათლება": "education",
    "ჰუმანიტარული მეცნიერებები და ხელოვნება": "humanities",
    "სოციალური მეცნიერებები, ბიზნესი და სამართალი": "social_business_law",
    "მეცნიერება": "science",
    "საინჟინრო, დამამუშავებელი და სამშენებლო დარგები": "engineering",
    "სოფლის მეურნეობა": "agriculture",
    "ჯანდაცვა და სოციალური უზრუნველყოფა": "health",
    "მომსახურება": "services",
}

FIELD_KA = {
    "education": "განათლება",
    "humanities": "ჰუმანიტარული მეცნიერებები და ხელოვნება",
    "social_business_law": "სოციალური მეცნიერებები, ბიზნესი და სამართალი",
    "science": "საბუნებისმეტყველო მეცნიერება",
    "engineering": "საინჟინრო და ტექნოლოგია",
    "agriculture": "სოფლის მეურნეობა",
    "health": "ჯანდაცვა და სოციალური უზრუნველყოფა",
    "services": "მომსახურება",
}


def find_year_cols(ws):
    cols = {}
    for i, cell in enumerate(ws[2]):
        if cell.value:
            for y in TARGET_YEARS:
                if y in str(cell.value):
                    cols[i] = y
    return cols


def parse_sheet(ws):
    col_years = find_year_cols(ws)
    result = {y: {} for y in TARGET_YEARS}

    # Section detection: only parse rows 4-13 (total = state + private)
    # Skip female/male subsets, skip state-only and private-only blocks
    in_total_block = False
    skip_next = False

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if not row[0]:
            continue
        label = str(row[0]).strip()

        # Detect section headers
        if "სახელმწიფო და კერძო" in label:
            in_total_block = True
            skip_next = False
            continue
        if "სახელმწიფო" in label and "კერძო" not in label:
            in_total_block = False
            continue
        if "კერძო" in label and "სახელმწიფო" not in label and len(label) < 20:
            in_total_block = False
            continue

        # Skip gender subsets
        if "მათ შორის ქალი" in label or "მათ შორის კაცი" in label:
            skip_next = True
            continue
        if label == "მათ შორის:":
            continue
        if skip_next and label in FIELD_MAP:
            # check if we're back to main section
            continue
        if "მიღებულთა რაოდენობა" in label or "კურსდამთავრებულ" in label:
            skip_next = False
            continue

        if not in_total_block:
            continue

        for ka, en in FIELD_MAP.items():
            if label.startswith(ka.strip()):
                for col_i, year in col_years.items():
                    val = row[col_i]
                    if val is None:
                        continue
                    try:
                        val = int(str(val).replace(" ", "").replace("\xa0", ""))
                    except:
                        continue
                    result[year][en] = val
                break

    return result


wb = openpyxl.load_workbook(os.path.join(RAW, "01_admissions_by_program.xlsx"))

# Parse admissions
admissions = parse_sheet(wb["მიღება"])

# Parse graduates
graduates = parse_sheet(wb["კურსდამთავრებულები"])

# Parse enrollment (total enrolled at start of year)
enrollment = parse_sheet(wb["რიცხოვნობა"])

# Calculate trends: compare 2022-2023 to 2024-2025
trends = {}
for field in FIELD_MAP.values():
    v2022 = admissions["2022-2023"].get(field)
    v2024 = admissions["2024-2025"].get(field)
    if v2022 and v2024:
        pct = ((v2024 - v2022) / v2022) * 100
        if pct >= 8:
            trend = "growing"
        elif pct <= -8:
            trend = "declining"
        else:
            trend = "stable"
        trends[field] = {"trend": trend, "change_pct": round(pct, 1)}

output = {
    "source": "Statistics Georgia (GeoStat)",
    "url": "https://www.geostat.ge/ka/modules/categories/61/umaghlesi-ganatleba",
    "note": "Only real data from GeoStat. No fabricated values.",
    "years": TARGET_YEARS,
    "admissions": admissions,
    "enrollment": enrollment,
    "graduates": graduates,
    "trends": trends,
}

out_path = os.path.join(OUT, "geostat_v2.json")
with open(out_path, "w", encoding="utf-8") as f:
    json.dump(output, f, ensure_ascii=False, indent=2)

print("=== ADMISSIONS (state + private total) ===")
for y in TARGET_YEARS:
    total = sum(admissions[y].values())
    print(f"{y}: total={total:,}  |  {admissions[y]}")

print("\n=== GRADUATES ===")
for y in TARGET_YEARS:
    total = sum(graduates[y].values())
    print(f"{y}: total={total:,}  |  {graduates[y]}")

print("\n=== TRENDS (2022→2024) ===")
for f, t in trends.items():
    print(f"  {f}: {t['trend']} ({t['change_pct']:+.1f}%)")

print(f"\nSaved: {out_path}")
